"""Read an FOI agency xlsx into raw sheet dicts. Pure openpyxl."""
from pathlib import Path
from openpyxl import load_workbook

def read_sheets(path: Path) -> dict[str, list[list]]:
    """Return {sheet_name: [rows as lists]} with formulas resolved to values."""
    wb = load_workbook(path, data_only=True, read_only=True)
    out = {}
    for name in wb.sheetnames:
        rows = []
        for row in wb[name].iter_rows(values_only=True):
            rows.append(list(row))
        out[name] = rows
    return out
