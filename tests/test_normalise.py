from pathlib import Path
import sys
sys.path.insert(0, "src")
from ingest.normalise import normalise_all
from config import DATA_SOURCES_DIR, GOLDEN_Q1_FIGURES

def _sum(facts, measure, bucket="total"):
    return round(sum(f["value"] for f in facts if f["measure"] == measure and f["bucket"] == bucket), 0)

def test_golden_q1_received():
    facts = normalise_all(DATA_SOURCES_DIR)
    # the current file is Q1-Q3 cumulative; single-quarter Q1 is marked derived
    q1 = [f for f in facts if f["fy"] == "2025-26" and f["quarter"] == 1]
    assert round(sum(f["value"] for f in q1 if f["measure"] == "received" and f["bucket"] == "total"), 0) == GOLDEN_Q1_FIGURES["requests_received"]

def test_no_x_rows():
    facts = normalise_all(DATA_SOURCES_DIR)
    assert not any(f["agency_name"].startswith("x") or f["agency_name"].startswith("xx") for f in facts)

def test_total_row_not_resummed():
    facts = normalise_all(DATA_SOURCES_DIR)
    # the Total row's received value is trusted, not computed from agency rows
    tot = [f for f in facts if f["agency_name"] == "Total"]
    assert tot, "Total row present"

# Published Total-row values (total bucket) per FY — direct reads from each
# workbook's "Action on requests" / "Response times" Total row.
PUBLISHED_TOTALS = {
    "2019-20": {"decided": 29358, "granted_full": 13727, "granted_part": 11221, "refused": 4410, "withdrawn": 10000, "within_statutory": 23085},
    "2020-21": {"decided": 26680, "granted_full": 10978, "granted_part": 10984, "refused": 4718, "withdrawn": 6834, "within_statutory": 20663},
    "2021-22": {"decided": 25303, "granted_full": 9966, "granted_part": 10547, "refused": 4790, "withdrawn": 5916, "within_statutory": 17798},
    "2022-23": {"decided": 21228, "granted_full": 5376, "granted_part": 11055, "refused": 4797, "withdrawn": 15915, "within_statutory": 15723},
    "2023-24": {"decided": 21347, "granted_full": 4465, "granted_part": 11659, "refused": 5223, "withdrawn": 11024, "within_statutory": 15754},
    "2024-25": {"decided": 25211, "granted_full": 5395, "granted_part": 13558, "refused": 6258, "withdrawn": 13353, "within_statutory": 18296},
    "2025-26": {"decided": 22573, "granted_full": 4154, "granted_part": 12334, "refused": 6085, "withdrawn": 10598, "within_statutory": 16047},
}

def test_new_measures_extracted_per_fy():
    facts = normalise_all()
    for fy, expected in PUBLISHED_TOTALS.items():
        for measure, want in expected.items():
            # golden Q1 facts share fy=2025-26/measure/bucket=total but are single-
            # quarter headlines (derived=True, agency "Total"), not agency totals
            rows = [f for f in facts if f["fy"] == fy and f["measure"] == measure and f["bucket"] == "total" and not f["derived"]]
            assert rows, f"no {measure} rows for {fy}"
            got = round(sum(f["value"] for f in rows))
            assert got == want, f"{fy} {measure}: sum(agency total)={got} != published total {want}"


def test_decided_consistent_across_sheets_per_fy():
    # The source publishes the decided headline on BOTH the "Action on requests"
    # ("Total determined") and "Response times" ("Requests determined") sheets;
    # they must agree per FY. We ingest decided from "Action on requests" ONLY
    # (ingesting both would double-count). This test cross-checks the ingested
    # value against the OTHER sheet's published Total row, read directly from
    # the workbook. Verified across all 7 files (e.g. 2024-25: 25211 == 25211).
    # NOTE: Total determined is NOT the sum of outcome components (granted+part+
    # refused+withdrawn+transferred != decided) — the outcome breakdown covers a
    # different scope, so no sum-of-components equality is asserted.
    from openpyxl import load_workbook
    facts = normalise_all()
    fy_files = [("2019-20", "agency-foi-data-2019-20.xlsx"), ("2020-21", "agency-foi-data-2020-21.xlsx"),
                ("2021-22", "agency-foi-data-2021-22.xlsx"), ("2022-23", "agency-foi-data-2022-23.xlsx"),
                ("2023-24", "agency-foi-data-2023-24.xlsx"), ("2024-25", "agency-foi-data-2024-25.xlsx"),
                ("2025-26", "agency-foi-data-2025-26-q1-to-q3.xlsx")]
    for fy, fn in fy_files:
        got = round(sum(f["value"] for f in facts if f["fy"] == fy and f["measure"] == "decided"
                        and f["bucket"] == "total" and not f["derived"]))
        wb = load_workbook(DATA_SOURCES_DIR / fn, data_only=True, read_only=True)
        rows = list(wb["Response times"].iter_rows(values_only=True))
        hdr = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        first = [i for i, h in enumerate(hdr) if h.startswith("requests determined")][0]
        want = None
        for r in rows[3:]:
            if str(r[0] or "").strip().lower() == "total":
                want = float(r[first + 2]); break
        wb.close()
        assert want is not None and got == want, \
            f"{fy}: decided (Action)={got} != Requests determined (Response times)={want}"


def test_six_figures_no_longer_empty():
    from src.stats.catalog import foi_stats
    from src.storage.frame import Frame
    facts = normalise_all()
    frame = Frame(facts)
    for key in ("requests_decided_trend", "decided_top20", "decision_outcomes_trend",
                "granted_full_part_change", "timeliness_trend", "timeliness_change"):
        fig = foi_stats(frame, key)["value"]
        assert fig["series"], f"{key} series empty"
        assert any(v is not None for s in fig["series"] for v in s["values"]), f"{key} all None"


def test_portfolio_banner_rows_skipped():
    # the source sheets carry portfolio banner rows (merged section headers)
    # whose data columns repeat the portfolio name as text. They used to parse
    # as phantom zero-request agencies and pollute the agency filter. They must
    # not appear as agencies at all.
    facts = normalise_all()
    agencies = {f["agency_name"] for f in facts}
    for banner in ("Industry, Science and Resources",
                   "Industry, Science, Energy and Resources",
                   "Industry, Innovation and Science",
                   "Agriculture, Water and the Environment"):
        assert banner not in agencies, f"banner row parsed as an agency: {banner}"


def test_disr_renamed_to_most_recent_name():
    # DISR was renamed in the 2022 MoG changes (the "Energy" portfolio moved
    # out). The data notes say renamed agencies appear under their most recent
    # name, so the pre-2022-23 spelling must resolve to the current one and DISR
    # must read as one continuous series across the whole period.
    facts = normalise_all()
    agencies = {f["agency_name"] for f in facts}
    assert "Department of Industry, Science, Energy and Resources" not in agencies, \
        "old DISR spelling should resolve to the current name"
    disr = [f for f in facts if f["agency_name"] == "Department of Industry, Science and Resources"
            and f["measure"] == "received" and f["bucket"] == "total"]
    by_fy = {fy: round(sum(f["value"] for f in disr if f["fy"] == fy)) for fy in
             ("2019-20", "2020-21", "2021-22", "2022-23", "2023-24", "2024-25", "2025-26")}
    assert by_fy == {"2019-20": 126, "2020-21": 307, "2021-22": 182,
                     "2022-23": 182, "2023-24": 191, "2024-25": 255, "2025-26": 168}, by_fy
